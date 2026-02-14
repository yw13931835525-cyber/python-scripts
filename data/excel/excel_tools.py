#!/usr/bin/env python3
"""
Excel自动化工具
支持合并/拆分/汇总/格式转换
"""

import pandas as pd
import openpyxl
from typing import List, Dict
from dataclasses import dataclass


@dataclass
class ExcelConfig:
    """Excel配置"""
    engine: str = 'openpyxl'
    header_row: int = 0


class ExcelTools:
    """Excel工具类"""
    
    def __init__(self, config: ExcelConfig = None):
        self.config = config or ExcelConfig()
    
    def read(self, filepath: str) -> pd.DataFrame:
        """读取Excel"""
        return pd.read_excel(
            filepath,
            engine=self.config.engine,
            header=self.config.header_row
        )
    
    def write(self, df: pd.DataFrame, filepath: str, sheet_name: str = 'Sheet1'):
        """写入Excel"""
        df.to_excel(
            filepath,
            sheet_name=sheet_name,
            engine=self.config.engine,
            index=False
        )
    
    def merge_files(self, files: List[str], output: str, add_source: bool = True):
        """合并多个Excel文件"""
        dfs = []
        
        for f in files:
            if pd.io.excel.exists_excel(f):
                df = self.read(f)
                if add_source:
                    df['source_file'] = f
                dfs.append(df)
        
        if dfs:
            merged = pd.concat(dfs, ignore_index=True)
            self.write(merged, output)
            print(f"已合并 {len(dfs)} 个文件到 {output}")
    
    def split_by_column(self, filepath: str, column: str, output_dir: str):
        """按列值拆分Excel"""
        df = self.read(filepath)
        
        import os
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        for value in df[column].unique():
            subset = df[df[column] == value]
            filename = f"{value}.xlsx"
            filepath = f"{output_dir}/{filename}"
            self.write(subset, filepath)
            print(f"已保存 {filepath}")
    
    def add_formula(self, filepath: str, column: str, formula: str, result_column: str):
        """添加公式列"""
        df = self.read(filepath)
        df[result_column] = df.eval(formula)
        self.write(df, filepath)
        print(f"已添加公式列: {result_column}")
    
    def summarize(self, filepath: str, group_by: str, columns: List[str], output: str):
        """汇总统计"""
        df = self.read(filepath)
        
        agg_dict = {col: ['sum', 'mean', 'count'] for col in columns}
        summary = df.groupby(group_by).agg(agg_dict)
        
        # 扁平化列名
        summary.columns = ['_'.join(col).strip() for col in summary.columns.values]
        self.write(summary.reset_index(), output)
        print(f"已保存汇总到 {output}")
    
    def format_cells(self, filepath: str, column: str, format_type: str = 'currency'):
        """格式化单元格"""
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # 查找列索引
        header = [cell.value for cell in ws[1]]
        if column in header:
            col_idx = header.index(column) + 1
            
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                
                if format_type == 'currency':
                    cell.number_format = '¥#,##0.00'
                elif format_type == 'percent':
                    cell.number_format = '0.00%'
                elif format_type == 'date':
                    cell.number_format = 'YYYY-MM-DD'
        
        wb.save(filepath)
        print(f"已格式化 {column} 列")
    
    def remove_duplicates(self, filepath: str, subset: List[str] = None):
        """去除重复行"""
        df = self.read(filepath)
        original_len = len(df)
        
        df = df.drop_duplicates(subset=subset)
        
        removed = original_len - len(df)
        self.write(df, filepath)
        print(f"已去除 {removed} 个重复行")
    
    def fill_empty(self, filepath: str, column: str, value: str = '0'):
        """填充空值"""
        df = self.read(filepath)
        df[column] = df[column].fillna(value)
        self.write(df, filepath)
        print(f"已填充 {column} 列的空值")
    
    def convert_format(self, input_file: str, output_file: str, output_format: str = 'csv'):
        """格式转换"""
        df = self.read(input_file)
        
        if output_format == 'csv':
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
        elif output_format == 'json':
            df.to_json(output_file, orient='records', force_ascii=False, indent=2)
        elif output_format == 'excel':
            self.write(df, output_file)
        
        print(f"已转换为 {output_format} 格式: {output_file}")


# 示例使用
if __name__ == "__main__":
    excel = ExcelTools()
    
    # 合并文件
    # excel.merge_files(["a.xlsx", "b.xlsx"], "merged.xlsx")
    
    # 拆分文件
    # excel.split_by_column("data.xlsx", "category", "output/")
    
    # 汇总统计
    # excel.summarize("sales.xlsx", "产品", ["销售额", "数量"], "summary.xlsx")
    
    # 格式转换
    # excel.convert_format("data.xlsx", "data.csv", "csv")
    
    print("Excel工具已就绪！")
